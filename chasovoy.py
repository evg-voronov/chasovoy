import cv2
import openpyxl
import winsound
import time
import os
import random
from mtcnn.mtcnn import MTCNN


def audio_message(path, here):
    if here == 0:
        pass
    elif time.strftime('%H', time.gmtime(time.time()-here)) == "01":
        wav = random.choice(os.listdir(path + r'\audio'))
        winsound.PlaySound(path + r'\audio\\' + wav, winsound.SND_FILENAME)


def show_face_frame():
    if faces:
        bounding_box = faces[0]['box']
        keypoints = faces[0]['keypoints']
        cv2.rectangle(img,
                      (bounding_box[0], bounding_box[1]),
                      (bounding_box[0] + bounding_box[2], bounding_box[1] + bounding_box[3]),
                      (0, 0, 255),
                      2)
        cv2.circle(img, (keypoints['left_eye']), 3, (0, 0, 255), 2)
        cv2.circle(img, (keypoints['right_eye']), 2, (0, 0, 255), 2)
        cv2.circle(img, (keypoints['nose']), 3, (0, 0, 255), 2)
        cv2.circle(img, (keypoints['mouth_left']), 3, (0, 0, 255), 2)
        cv2.circle(img, (keypoints['mouth_right']), 3, (0, 0, 255), 2)
    cv2.imshow('frame', img)


def write_excel(path_excel, time_start):
    wb = openpyxl.load_workbook(path_excel + r'\dnevnik.xlsx')
    today = time.strftime("%d.%m.%Y", time.localtime(time.time()))
    if today not in wb.sheetnames:  # если листа с сегодняшней датой еще не существует, то создаем его
        wb.create_sheet(title=today)
        sheet = wb[today]
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 31
        sheet['A1'] = 'Начал работать за ПК'
        sheet['B1'] = 'Время работы за ПК'
        sheet['C1'] = 'Закончил работать'
        sheet['D1'] = 'Сел за компьютер:'
        sheet['D2'] = 'Всего провели за ПК времени:'
        sheet = wb[today]
        row = 2
        all_time = 0
    else:  # если лист уже создан извлекаем информацию
        sheet = wb[today]
        row = sheet['E1'].value  # номер строки в excel документе
        all_time = sheet['E2'].value  # общее время проведенное за ПК за один день
        all_time = all_time.split(':')
        all_time = int(all_time[0]) * 3600 + int(all_time[1]) * 60 + int(all_time[2])  # время переводим в секунды
        row = row + 2  # плюсуем шапку таблицы и новую строку

    sheet['A' + str(row)] = time.strftime("%H:%M:%S", time.localtime(time_start))  # записываем время когда сели за пк
    sheet['C' + str(row)] = time.strftime("%H:%M:%S", time.localtime(time.time()))  # когда вышли из за пк
    seconds = time.time() - time_start  # время которое мы провели за компьютером
    sheet['B' + str(row)] = time.strftime('%H:%M:%S', time.gmtime(seconds))
    sheet['E1'] = row - 1  # вычитаем шапку таблицу
    all_time = all_time + seconds
    all_time = time.strftime('%H:%M:%S', time.gmtime(all_time))
    sheet['E2'] = all_time
    # заполняем первый лист
    sheet_0 = wb.worksheets[0]
    sheet_0['A' + str(len(wb.sheetnames))] = today
    sheet_0['B' + str(len(wb.sheetnames))] = all_time
    while True:  # закрываем файл
        try:
            wb.save(path + r'\dnevnik.xlsx')
        except PermissionError:
            input('Сначала закройте excel файл и подтвердите нажав enter')
        else:
            break


cap = cv2.VideoCapture(0)
path = r'C:\Users\VoronovEV\YandexDisk-euge.voronov@yandex.ru\Manual\Project\chasovoy'  # Укажите свой абсолютный путь

here, not_here = 0, 0  # секундомеры
time_here, time_not_here = 0, 0  # количество времени
switch = True  # переключатель определяет есть лицо в кадре или нет
min_time = 120  # минимально время которое не учитывается (в секундах)

detector = MTCNN()

try:
    while cap.isOpened():
        _, img = cap.read()
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        faces = detector.detect_faces(img)
        show_face_frame()
        audio_message(path, here)

        if faces and switch:
            not_here = 0
            switch = False
            if here == 0:
                here = time.time()
                print('вы сели за компьютер в ' + time.strftime("%H:%M:%S", time.localtime(here)))
        elif not faces and not switch:
            not_here = time.time()
            switch = True

        if not_here != 0:
            time_not_here = time.time() - not_here
        if here != 0 and not switch:
            time_here = time.time() - here

        if time_here > min_time and time_not_here > min_time:  # делаем запись если провели за ПК больше min_time
            write_excel(path, here)
            print('вы провели за компьютером ' + time.strftime('%H:%M:%S', time.gmtime(time.time() - here)))
            time_here, time_not_here = 0, 0
            here, not_here = 0, 0
        elif time_not_here > min_time > time_here:  # если за ПК провели меньше min_time, то ничего не записываем
            print('запись не произведена')
            time_here, time_not_here = 0, 0
            here, not_here = 0, 0

        cv2.waitKey(200)
except KeyboardInterrupt:
    if time_here > min_time:  # делаем запись если провели за ПК больше минуты
        write_excel(path, here)
    cap.release()
    cv2.destroyAllWindows()
