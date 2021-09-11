import time
import cv2
import openpyxl
from mtcnn.mtcnn import MTCNN


class Excel:
    def __init__(self, today):
        self.today = today
        self.wb = openpyxl.load_workbook('log.xlsx')
        if today not in self.wb.sheetnames:
            self.new_list()
        else:
            self.extract_info()

    def new_list(self):
        self.wb.create_sheet(title=self.today)
        self.sheet = self.wb[self.today]
        self.sheet.column_dimensions['A'].width = 20
        self.sheet.column_dimensions['B'].width = 20
        self.sheet.column_dimensions['C'].width = 20
        self.sheet.column_dimensions['D'].width = 31
        self.sheet['A1'] = 'Начал работать за ПК'
        self.sheet['B1'] = 'Время работы за ПК'
        self.sheet['C1'] = 'Закончил работать'
        self.sheet['D1'] = 'Сел за компьютер:'
        self.sheet['D2'] = 'Всего провели за ПК времени:'
        self.row = 2
        self.all_time = 0

    def extract_info(self):
        self.sheet = self.wb[self.today]
        self.row = self.sheet['E1'].value  # line number in excel document
        self.all_time = self.sheet['E2'].value  # total time spent on PC in one day
        self.all_time = self.all_time.split(':')
        self.row = self.row + 2  # add the table header and a new line
        # we convert time into seconds
        self.all_time = int(self.all_time[0]) * 3600 + int(self.all_time[1]) * 60 + int(self.all_time[2])

    def record(self, start_work, time_break):
        # we write down the time when we sat down at the PC
        self.sheet['A' + str(self.row)] = time.strftime("%H:%M:%S",
                                                        time.localtime(start_work))
        self.sheet['C' + str(self.row)] = time.strftime("%H:%M:%S", time.localtime(
            time.time() - time_break))  # when left from the PC
        seconds = time.time() - start_work - time_break  # the time we spent at the computer
        self.sheet['B' + str(self.row)] = time.strftime('%H:%M:%S', time.gmtime(seconds))
        self.sheet['E1'] = self.row - 1  # subtract the table header
        self.all_time = self.all_time + seconds
        self.sheet['E2'] = time.strftime('%H:%M:%S', time.gmtime(self.all_time))
        # fill in the first sheet
        sheet_0 = self.wb.worksheets[0]
        sheet_0['A' + str(len(self.wb.sheetnames))] = self.today
        sheet_0['B' + str(len(self.wb.sheetnames))] = time.strftime('%H:%M:%S', time.gmtime(self.all_time))
        self.row += 1

    def close_file(self):
        open = True
        while open:  # close the file
            try:
                self.wb.save('log.xlsx')
                open = False
            except PermissionError:
                input('Сначала закройте excel файл и подтвердите нажав enter')


class SessionWork:
    def __init__(self):
        self.start_work = 0
        self.start_break = 0
        self.time_break = 0
        self.time_work = 0

    def sit(self):
        if self.start_work == 0:
            self.start_work = time.time()
        self.start_break = 0
        self.time_break = 0

    def stand_up(self):
        if self.start_break == 0 and self.start_work != 0:
            self.start_break = time.time()

    def status(self):
        if self.start_work != 0 and self.start_break == 0:
            self.time_work = time.time() - self.start_work
        if self.start_break != 0:
            self.time_break = time.time() - self.start_break

    def show_faces(self, img, faces):
        if faces:
            bounding_box = faces[0]['box']
            cv2.rectangle(img,
                          (bounding_box[0], bounding_box[1]),
                          (bounding_box[0] + bounding_box[2], bounding_box[1] + bounding_box[3]),
                          (0, 0, 255),
                          2)
        cv2.imshow("I", img)


cap = cv2.VideoCapture(0)
detector = MTCNN()
min_time_work = 1800  # seconds. The minimum time after which the recording in Excel
min_time_break = 300  # seconds. The minimum break time after which the recording in Excel
I = SessionWork()
today = time.strftime("%d.%m.%Y", time.localtime(time.time()))
excel_sheet = Excel(today)

while cap.isOpened():
    _, img = cap.read()
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    faces = detector.detect_faces(img)
    if faces:
        I.sit()
    else:
        I.stand_up()
    I.status()
    print(f"start_work {I.start_work}")
    print(f"start_break {I.start_break}")
    print(f"time_work {I.time_work}")
    print(f"time_break {I.time_break}")
    print("-------------------")
    if I.time_work > min_time_work and I.time_break > min_time_break:
        excel_sheet.record(I.start_work, I.time_break)
        print("record")
        I.__init__()
        excel_sheet.close_file()
    #I.show_faces(img, faces)
    cv2.waitKey(800)

cap.release()
cv2.destroyAllWindows()
